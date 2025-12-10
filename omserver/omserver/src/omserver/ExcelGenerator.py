from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from flask import current_app
class ExcelGenerator:
    """
    Export batch simulation results to Excel file with multiple sheets.
    
    Args:
        batch_data: Dictionary containing batch simulation results
        
    Returns:
        Filename of the created Excel file
    """
    def __init__(self, batch_data):
        self.batch_data = batch_data
    def export_batch_to_excel(self):
        """
        Export batch simulation results to Excel file with multiple sheets.
        
        Args:
            batch_data: Dictionary containing batch simulation results
            
        Returns:
            Filename of the created Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sheet 1: Batch Summary
        ws_batch = wb.create_sheet("Batch_Summary", 0)
        self._create_batch_summary_sheet(ws_batch, self.batch_data)
        
        # Sheet 2: Iterations Overview
        ws_overview = wb.create_sheet("Iterations_Overview", 1)
        self._create_iterations_overview_sheet(ws_overview, self.batch_data)
        
        # Sheets 3+: Individual iteration sheets
        iterations = self.batch_data.get("batch_sim_res_collection", [])
        for idx, iteration in enumerate(iterations):
            sheet_name = self._generate_iteration_sheet_name(iteration, idx)
            ws_iter = wb.create_sheet(sheet_name, idx + 2)
            self._create_iteration_detail_sheet(ws_iter, iteration)
        
        # Save Excel file
        safe_title = self._sanitize_filename(self.batch_data.get("batch_sim_title", "batch"))
        filename = f"{safe_title}.xlsx"
        output_path = Path(current_app.instance_path) / filename
        wb.save(output_path)
        
        return filename
    
    def _create_batch_summary_sheet(self, ws, batch_data):
        """Create the batch summary sheet with basic batch information."""
        ws.append(["Batch Simulation Summary"])
        ws.append([])  # Empty row
        
        # Add batch information
        ws.append(["Batch Title", batch_data.get("batch_sim_title", "N/A")])
        ws.append(["Timestamp", batch_data.get("batch_sim_time_stamp", "N/A")])
        ws.append(["Vessel Name", batch_data.get("vessel_name", "N/A")])
        ws.append(["Batch Size", batch_data.get("batch_size", 0)])
        ws.append(["Actual Iterations", len(batch_data.get("batch_sim_res_collection", []))])
        
        # Auto-size columns
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_iterations_overview_sheet(self,ws, batch_data):
        """Create the iterations overview sheet with summary of all iterations."""
        iterations = batch_data.get("batch_sim_res_collection", [])
        
        # Header row
        headers = [
            "Iteration ID",
            "Iteration Name",
            "Gen1 Name",
            "Gen1 Lower (%)",
            "Gen1 Upper (%)",
            "Gen2 Name",
            "Gen2 Lower (%)",
            "Gen2 Upper (%)",
            "Gen3 Name",
            "Gen3 Lower (%)",
            "Gen3 Upper (%)",
            "Battery Name",
            "Battery Count",
            "Battery Rated Power (kW)",
            "Total Gen1 Energy (kWh)",
            "Total Gen2 Energy (kWh)",
            "Total Gen3 Energy (kWh)",
            "Total Battery Charging Energy (kWh)",
            "Total Battery Discharging Energy (kWh)",
            "Diesel Usage (Ton)",
            "Methanol Usage (Ton)",
            "Hydrogen Usage (Ton)",
            "Total Energy Demand (kWh)",
            "Total Energy Supplied (kWh)",
            "Total Energy Wasted (kWh)",
            "Energy Balance Check",
            "Wasted Energy Significance Check",
            "Limitation Check",
            "CO2 Emission (Ton)",
            "Penalty (EUR)",
            "Gen1 Cost (EUR)",
            "Gen2 Cost (EUR)",
            "Gen3 Cost (EUR)",
            "Battery Cost (EUR)",
            "Battery Cycle Limit",
            "Battery Total Mass (kg)",
            "Battery Total Volume (m^3)",
            "Gen1 Volume (m^3)",
            "Gen2 Volume (m^3)",
            "Gen3 Volume (m^3)",
            "Gen1 Mass (KG)",
            "Gen2 Mass (KG)",
            "Gen3 Mass (KG)"
        ]
        ws.append(headers)
        
        # Data rows
        for iteration in iterations:
            sime_name = iteration.get("sim_name", "N/A")
            sequence = iteration.get("sequence", "N/A")
            battery_name = iteration.get("battery_name", "N/A")
            battery_count = iteration.get("battery_count", "N/A")
            opt_zone = iteration.get("optimalZone", [0, 0, 0, 0, 0, 0])
            gen_costs = iteration.get("Gen Costs", [0,0,0])
            gen_mass = iteration.get("Gen Mass", [0,0,0])
            gen_volume = iteration.get("Gen Volumes", [0,0,0])
            battery_spec = iteration.get("Battery Specs", [0,0,0])
            energy_balance = "True" if iteration.get("Total Energy Wasted (kWh)") == 0 else "False"
            wastered_energy_significance_check = "True" if iteration.get("Total Energy Wasted (kWh)") >  10  else "False"
            limitation_check  = "Use Less Engine" if iteration.get("Total Energy Wasted (kWh)") >  10 else "Use More Battery" if  iteration.get("Total Energy Wasted (kWh)") < -10 else "None"
            # Parse sequence to extract generator names
            gen_names =self._parse_generator_names_from_sequence(sequence)
            
            row = [
                iteration.get("iteration_id", "N/A"),
                sime_name,
                gen_names[0], opt_zone[0] if len(opt_zone) > 0 else 0, opt_zone[1] if len(opt_zone) > 1 else 0,
                gen_names[1], opt_zone[2] if len(opt_zone) > 2 else 0, opt_zone[3] if len(opt_zone) > 3 else 0,
                gen_names[2], opt_zone[4] if len(opt_zone) > 4 else 0, opt_zone[5] if len(opt_zone) > 5 else 0,
                battery_name,
                battery_count,
                battery_spec[0],
                iteration.get("Total Gen1 Energy (kWh)", 0),
                iteration.get("Total Gen2 Energy (kWh)", 0),
                iteration.get("Total Gen3 Energy (kWh)", 0),
                iteration.get("Total Battery Charging Energy (kWh)", 0),
                iteration.get("Total Battery Discharging Energy (kWh)", 0),
                iteration.get("diesel_usage (Ton)", 0),
                iteration.get("meth_usage (Ton)", 0),
                iteration.get("hydrogen_usage (Ton)", 0),
                iteration.get("Total Energy Deamand (kWh)", 0),
                iteration.get("Total Energy Supplied (kWh)", 0),
                iteration.get("Total Energy Wasted (kWh)", 0),
                energy_balance,
                wastered_energy_significance_check,
                limitation_check,
                iteration.get("CO2_emission (Ton)", 0),
                iteration.get("penalty (EUR)", 0),
                gen_costs[0],
                gen_costs[1],
                gen_costs[2],
                battery_spec[1],
                battery_spec[2],
                battery_spec[4],
                battery_spec[3],
                gen_volume[0],
                gen_volume[1],
                gen_volume[2],
                gen_mass[0],
                gen_mass[1],
                gen_mass[2]
            ]
            ws.append(row)
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _generate_iteration_sheet_name(self, iteration, idx):
        """Generate a unique sheet name for an iteration (max 31 characters)."""
        iter_id = iteration.get("iteration_id", idx)
        
        # Try to create a short descriptive name
        sim_name = iteration.get("sim_name", "")
        
        # Extract a short identifier from sim_name (first engine name)
        short_name = ""
        if "Slot1:" in sim_name:
            parts = sim_name.split("|")
            if parts:
                slot1 = parts[0].replace("Slot1:", "").strip()
                # Take first word or first 10 chars
                short_name = slot1.split()[0][:10] if slot1 else ""
        
        # Create name: "Iter_123_EngineName"
        if short_name:
            sheet_name = f"Iter_{iter_id}_{short_name}"
        else:
            sheet_name = f"Iteration_{iter_id}"
        
        # Ensure it's under 31 characters (Excel limit)
        if len(sheet_name) > 31:
            sheet_name = f"Iter_{iter_id}"
        
        return sheet_name
    
    def _create_iteration_detail_sheet(self, ws, iteration):
        """Create detailed sheet for a single iteration."""
        # Section 1: Basic Information
        ws.append(["Iteration Details"])
        ws.append([])
        
        ws.append(["Simulation Name", iteration.get("sim_name", "N/A")])
        ws.append(["Sequence", iteration.get("sequence", "N/A")])
        ws.append(["Iteration ID", iteration.get("iteration_id", "N/A")])
        
        # Optimal zones with labels
        opt_zone = iteration.get("optimalZone", [0, 0, 0, 0, 0, 0])
        ws.append(["Gen1 Optimal Zone", f"Lower: {opt_zone[0] if len(opt_zone) > 0 else 0}, Upper: {opt_zone[1] if len(opt_zone) > 1 else 0}"])
        ws.append(["Gen2 Optimal Zone", f"Lower: {opt_zone[2] if len(opt_zone) > 2 else 0}, Upper: {opt_zone[3] if len(opt_zone) > 3 else 0}"])
        ws.append(["Gen3 Optimal Zone", f"Lower: {opt_zone[4] if len(opt_zone) > 4 else 0}, Upper: {opt_zone[5] if len(opt_zone) > 5 else 0}"])
        
        ws.append([])
        ws.append(["Energy Summary"])
        ws.append(["Total Energy Demand (kWh)", iteration.get("Total Energy Deamand (kWh)", 0)])
        ws.append(["Total Energy Supplied (kWh)", iteration.get("Total Energy Supplied (kWh)", 0)])
        ws.append(["Total Energy Wasted (kWh)", iteration.get("Total Energy Wasted (kWh)", 0)])
        
        ws.append([])
        ws.append(["Fuel Consumption"])
        ws.append(["Diesel Usage (Ton)", iteration.get("diesel_usage (Ton)", 0)])
        ws.append(["Methanol Usage (Ton)", iteration.get("meth_usage (Ton)", 0)])
        ws.append(["Hydrogen Usage (Ton)", iteration.get("hydrogen_usage (Ton)", 0)])
        
        ws.append([])
        ws.append(["Emissions"])
        ws.append(["CO2 Emission (Ton)", iteration.get("CO2_emission (Ton)", 0)])
        ws.append(["Penalty (EUR)", iteration.get("penalty (EUR)", 0)])
        
        ws.append([])
        ws.append([])  # Space before time series data
        
        # Section 2: Time Series Data
        ws.append(["Time Series Data"])
        ws.append([])
        

        # Headers for time series
        time_series_headers = [
            "Time (h)",
            "Power Demand (kW)",
            "Gen 1 Power (kW)",
            "Gen 2 Power (kW)",
            "Gen 3 Power (kW)",
            "Gen 1 Energy (kWh)",
            "Gen 2 Energy (kWh)",
            "Gen 3 Energy (kWh)",
            "Battery SOC (%)",
            "Battery Discharge (KW)",
            "Battery Charge (KW)",
            "Battery Discharging Energy (kWh)",
            "Battery Charging Energy(kWh)",
            "Wasted Power (KW)",
            "Battery Measured Power (kW)",
            "Inefficient Performance (KW)",
            "Under Supply (KW)",
            "Power Balance Check",
            "Significant Check?",
        ]
        ws.append(time_series_headers)
        
        # Extract time series data
        time = iteration.get("time (h)", [])
        power_demand = iteration.get("power_demand (KW)", [])
        gen1_power = iteration.get("gen_1_power (KW)", [])
        gen2_power = iteration.get("gen_2_power (KW)", [])
        gen3_power = iteration.get("gen_3_power (KW)", [])
        gen1_energy = iteration.get("Gen1 Energy (kWh)", [])
        gen2_energy = iteration.get("Gen2 Energy (kWh)", [])
        gen3_energy = iteration.get("Gen3 Energy (kWh)", [])
        battery_soc = iteration.get("battery_soc (%)", [])
        battery_discharge = iteration.get("battery_discharge (KW)", [])
        battery_charge = iteration.get("battery_charge (KW)", [])
        battery_charge_energy = iteration.get("Battery Charging Energy(kWh)", [])
        battery_discharge_energy =iteration.get("Battery Discharging Energy(kWh)", [])
        wasted_power = iteration.get("Wasted Power (kW)", [])
        battery_measured = iteration.get("battery_measured_power (kW)", [])
        over_supply =  [w if w > 0 else 0 for w in wasted_power]
        under_supply = [w if w < 0 else 0 for w in wasted_power]
        power_balance_check = ["True" if w == 0 else "False" for w in wasted_power]
        signicant_check = ["True" if abs(w) > 5 else "False" for w in wasted_power]
    
        # Write time series data
        max_len = max(len(time), len(power_demand), len(gen1_power))
        for i in range(max_len):
            row = [
                time[i] if i < len(time) else "",
                power_demand[i] if i < len(power_demand) else "",
                gen1_power[i] if i < len(gen1_power) else "",
                gen2_power[i] if i < len(gen2_power) else "",
                gen3_power[i] if i < len(gen3_power) else "",
                gen1_energy[i] if i < len(gen1_energy) else "",
                gen2_energy[i] if i < len(gen2_energy) else "",
                gen3_energy[i] if i < len(gen3_energy) else "",
                battery_soc[i] if i < len(battery_soc) else "",
                battery_discharge[i] if i < len(battery_discharge) else "",
                battery_charge[i] if i < len(battery_charge) else "",
                battery_charge_energy[i] if i < len(battery_charge_energy) else "",
                battery_discharge_energy[i] if i < len(battery_discharge_energy) else "",
                wasted_power[i] if i < len(wasted_power) else "",
                battery_measured[i] if i < len(battery_measured) else "",
                over_supply[i] if i< len(over_supply) else "",
                under_supply[i] if i< len(under_supply) else "",
                power_balance_check[i] if i< len(power_balance_check) else "",
                signicant_check[i] if i <len(signicant_check) else "",
            ]
            ws.append(row)
        
        # Auto-size columns
        for col in range(1, len(time_series_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _sanitize_filename(self, filename):

        """Sanitize filename for filesystem compatibility."""
        # Remove invalid characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # Replace spaces with underscores
        filename = filename.replace(' ', '_')
        
        # Limit length
        if len(filename) > 200:
            filename = filename[:200]
        
        return filename 
    def _parse_generator_names_from_sequence(self, sequence):
        """Extract generator names from sequence string."""
        gen_names = ["None", "None", "None"]
        
        # Example sequence: "Gen1:[Fortuna Crane Engine] → Gen2:[...] → Gen3:[...] + Batt:[...]"
        try:
            parts = sequence.split("→")
            for i, part in enumerate(parts[:3]):  # Only first 3 generators
                if "[" in part and "]" in part:
                    name = part.split("[")[1].split("]")[0].strip()
                    gen_names[i] = name
        except:
            pass
        
        return gen_names

    
    