import base64
from itertools import product
import json
import  pandas as pd
from datetime import datetime
import shutil
import os
import numpy as np
from pathlib import Path
from multiprocessing  import Pool, Process
from flask_cors import CORS
from flask import Blueprint, current_app, request, jsonify
from omserver.EUEmissionCalculator import EUMarineEmissionCalculator
from omserver.ModelicaParamParser import ModelicaParamParser
from concurrent.futures import ProcessPoolExecutor, as_completed
from .OMCConnection import OMCConnection
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

bp = Blueprint("model", __name__, url_prefix="/model")
CORS(bp)
PROGRESS = {
    "running": False,
    "total": 0,
    "done": 0,
    "current": None,     
    "cancelled": False,
    "last_error": None,
}

ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\']")
MAX_SHEET_NAME_LEN = 31
@bp.route("/upload", methods=["POST"])
def upload():
    try:
        model_path = Path(current_app.instance_path) / f"{request.json['model_name']}.mo"
        # Ensure instance directory exists
        model_path.parent.mkdir(parents=True, exist_ok=True)
        # Write in binary mode to handle Unicode properly
        with open(model_path, "wb") as model_file:
            model_file.write(base64.b64decode(request.json['model_data']))
        print(f"Decoded model to {model_path}")
    except Exception as e:
        print(f"Error writing model file: {e}")
        return json.dumps({
            "name": request.json['model_name'],
            "status": "Error opening writable instance resource"
        })

    return json.dumps({
        "name": request.json['model_name'],
        "status": "Model written"
    })

@bp.route("/delete", methods=["POST"])
def delete():
    try:
        os.remove(
            f"{current_app.instance_path}/{request.json['model_name']}.mo")
    except FileNotFoundError:
        return json.dumps(
            {"status": f"Model {request.json['model_name']} not found"})

    return json.dumps(
        {"status": f"Deleted model {request.json['model_name']}"})

@bp.route("/simulate", methods=["POST"])
def simulate():
    omc = OMCConnection()
    omc.request(
        f"loadFile(\"{current_app.instance_path}/{request.json['model_name']}.mo\")"
    )   

    overrides = " -override "
    for override in request.json["overrides"]:
        overrides += f"{override['param']}={override['value']},"
    if overrides[-1] == ",":
        overrides = overrides[:-1]
    omc.request(
        f"simulate({request.json['model_name']}, outputFormat=\"csv\", startTime={request.json['start_time']}, stopTime={request.json['stop_time']}, simflags=\"-outputPath {current_app.instance_path}{overrides if len(overrides) > 11 else ''}\")"
    )

    result_string = "No result"
    with current_app.open_instance_resource(
            f"{request.json['model_name']}_res.csv", "r") as result:
        result_string = result.read()

    return json.dumps({
        "name": request.json['model_name'],
        "result": result_string
    })

    
@bp.route("/simulate_batch", methods=["POST"])
def simulate_batch():
    # Retrive Request
    data = request.get_json(force=True)
    model_name  = data["model_name"]
    start_time  = data["start_time"]
    stop_time   = data["stop_time"]
    combos      = data["list_of_config_combinations"]

    # Log General Batch Simulation Info
    print(f"[simulate_batch] model={model_name}, combos={len(combos)}, "
          f"start={start_time}, stop={stop_time}")

    # Connect OMC 
    omc = OMCConnection()
    
    
    # Tracking Progress, Update the Global Status tracker  
    PROGRESS.update({
        "running": True,
        "total": len(combos),
        "done": 0,
        "current": None,
        "cancelled": False,
        "last_error": None,
    })

    # Check if the result collection db exist, if not , make one
    # This only Applied to First Time 
    path_to_db = Path(current_app.instance_path) / "res_db.json"
    path_to_db.parent.mkdir(parents=True, exist_ok=True)

    # Form Current Time
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    # Form the current bastch simulation object
    temp_result_collection = {
        "batch_sim_title" :  f"{current_time}_{model_name}_vessel_name_place_holder",
        "batch_sim_time_stamp" : f"{current_time}",
        "vessel_name" : "dev_vessel",
        "batch_size" : int(len(combos)),
        "batch_sim_res_collection"  : []
    }

    # Looping Start
    try:
        # Looping Start
        for idx, cfg in enumerate(combos):

            
            # Update the BSFC &&  FCC for this iteration
            diesel_bsfc = cfg.get("instance").get("Diesel_Engine").get("engine_bsfc")
            meth_bsfc = cfg.get("instance").get("Meth_Engine").get("engine_bsfc")
            diesel_fcc = cfg.get("instance").get("Diesel_Engine").get("engine_fcc")
            meth_fcc = cfg.get("instance").get("Meth_Engine").get("engine_fcc")
            modelicaArrayParser = ModelicaParamParser(model_name, diesel_bsfc, meth_bsfc,diesel_fcc, meth_fcc)
            modelicaArrayParser.update_modelica_txt_formate()
        
            # Log Simulation Detail
            print(f"[simulate_batch] starting {idx+1}/{len(combos)}")
            print(f"[simulation detail] Diesel : {cfg.get("instance").get("Diesel_Engine_Count")} * {cfg.get("instance").get("Diesel_Engine").get("engine_name")}")
            print(f"[simulation detail] Meth : {cfg.get("instance").get("Meth_Count")} * {cfg.get("instance").get("Meth_Engine").get("engine_name")}")
            print(f"[simulation detail] Battery : {cfg.get("instance").get("Battery_Count")} * {cfg.get("instance").get("Battery").get("battery_name")}")
            print(f"[simulation detail] Diesel_BSFC : {diesel_bsfc}") 
            print(f"[simulation detail] Meth_BSFC : {meth_bsfc}")
            print(f"[simulation detail] Diesel_FCC : {diesel_fcc}")
            print(f"[simulation detail] Meth_FCC : {meth_fcc}") 

            
            
            #Form the total cost of the configuration
            cost  = float(cfg.get("instance").get("Diesel_Engine").get("engine_cost")) * int(cfg.get("instance").get("Diesel_Engine_Count"))+ float(cfg.get("instance").get("Meth_Engine").get("engine_cost")) * int(cfg.get("instance").get("Meth_Count"))+ float(cfg.get("instance").get("Battery").get("battery_cost"))* int(cfg.get("instance").get("Battery_Count"))
            print(f"Cost of  Setup : {cost}  £  ")

            #Form the max potential power output of the power train
            max_potential_gen = float(cfg.get("instance").get("Diesel_Engine").get("engine_p_max")) * float(cfg.get("instance").get("Diesel_Engine_Count")) + float(cfg.get("instance").get("Meth_Engine").get("engine_p_max")) * float(cfg.get("instance").get("Meth_Count"))+ float(cfg.get("instance").get("Battery").get("battery_max_charge_power"))* (1/1000)* float(cfg.get("instance").get("Battery_Count"))
            print(f"Max Possible Power {max_potential_gen} w")
            
            # Unpack the parameter list
            overrides_list = cfg.get("changed_parameters", [])
            ov = " -override " + ",".join(f"{o['param']}={o['value']}" for o in overrides_list)
            simflags = f'-outputPath {current_app.instance_path}{ov if len(ov) > 20 else ""}'
            path_to_mo = Path(current_app.instance_path) / f"{model_name}.mo"
            
            with path_to_mo.open("r", encoding="utf-8") as f:
                for lineno, line in enumerate(f, start=1):
                    if "BSFC_Curve[:, 2]" in line:
                        print(f"{lineno}: {line.strip()}")
                    if "Engine_Fuel_Consumption_Look_Up_Table_Diesle[:, 2]" in line:
                        print(f"{lineno}: {line.strip()}")
            # Make request to  OMC, set over the Model, Model Name, etc
            try:
                omc.request(f'loadFile("{current_app.instance_path}/{model_name}.mo")')  
                omc.request(
                    f'simulate({model_name}, '
                    f'outputFormat="csv", '
                    f'startTime={start_time}, '
                    f'stopTime={stop_time}, '
                    f'numberOfIntervals = 7500, '
                    f'simflags="{simflags}")'
                )

                # Log if the simulation has finished 
                print(f"[simulate_batch] ✓ finished {idx+1}")
                print(f"[simulate_batch] start processing")
                
                simName = f"{cfg.get("instance").get("Diesel_Engine_Count")} * {cfg.get("instance").get("Diesel_Engine").get("engine_name")}_{cfg.get("instance").get("Meth_Count")} * {cfg.get("instance").get("Meth_Engine").get("engine_name")}_{cfg.get("instance").get("Battery_Count")} * {cfg.get("instance").get("Battery").get("battery_name")}"
                simResult = process_simmultion_result(
                    int(idx), 
                    simName, 
                    f"{cfg.get("instance").get("Diesel_Engine").get("engine_name")}",
                    f"{cfg.get("instance").get("Diesel_Engine_Count")}",
                    f"{cfg.get("instance").get("Meth_Engine").get("engine_name")}",
                    f"{cfg.get("instance").get("Meth_Count")}",
                    f"{cfg.get("instance").get("Battery").get("battery_name")}",
                    f"{cfg.get("instance").get("Battery_Count")}",
                    cost,
                    max_potential_gen
                )
                temp_result_collection["batch_sim_res_collection"].append(simResult)

            except Exception as e:
                print(f"[simulate_batch] ✗ error at {idx+1}: {e}")
                raise
            # Log Simulation Done 
            PROGRESS["done"] = idx + 1
        # Loop ends, Check Status
        status = "cancelled" if PROGRESS["cancelled"] else "completed"
        print(f"[simulate_batch] finished with status={status}, " f"done={PROGRESS['done']}/{PROGRESS['total']}")
        
        # Load Result into the database JSON 
        try:
            append_batch_result(temp_result_collection)
            current_app.logger.info(
                "Saved batch: %s (size=%d)", temp_result_collection["batch_sim_title"], len(temp_result_collection["batch_sim_res_collection"])
            )
        except Exception as e: 
            current_app.logger.exception("[simulate_batch_end] ERROR: %s", e)
    except Exception as e:
        PROGRESS["last_error"] = str(e)
        print(f"[simulate_batch] ERROR: {e}")
        return jsonify({"status": "error", "error": str(e)}), 500
    finally:
        PROGRESS["running"] = False
        PROGRESS["current"] = None

        # Form a list of all avalible batch
        print("Saving finally done")
        avalible_batches = []
        return jsonify({
            "status": status, 
            "done": PROGRESS["done"], 
            "total": PROGRESS["total"],
            "avalible_batches" : avalible_batches
        })

@bp.route("/simulate_batch_mp", methods=["POST"])
def simulate_batch_mp():
    data = request.get_json(force=True)
    model_name  = data["model_name"]
    start_time  = data["start_time"]
    stop_time   = data["stop_time"]
    combos      = data["list_of_config_combinations"]

    print(f"[simulate_batch_mp] model={model_name}, combos={len(combos)}, start={start_time}, stop={stop_time}")

    instance_path = current_app.instance_path  # pass this into workers

    # Build batch record container
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    batch_record = {
        "batch_sim_title": f"{current_time}_{model_name}_vessel_name_place_holder",
        "batch_sim_time_stamp": current_time,
        "vessel_name": "dev_vessel",
        "batch_size": int(len(combos)),
        "batch_sim_res_collection": [],
    }

    results = []
    errors  = []

    # NOTE: Choose sensible worker count
    max_workers = min(os.cpu_count() or 2, 4)  # cap if OMC/CPU constrained

    # Submit all tasks first; then gather as they complete
    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futures = [
            ex.submit(single_simulation, model_name, start_time, stop_time, idx, cfg, instance_path)
            for idx, cfg in enumerate(combos)
        ]

        for fut in as_completed(futures):
            try:
                res = fut.result()
                results.append(res)
            except Exception as e:
                print("[simulate_batch_mp] worker error:", e)
                errors.append(str(e))

    # Attach the results and append to the JSON “DB”
    batch_record["batch_sim_res_collection"] = results

    try:
        append_batch_result(batch_record)  
        current_app.logger.info(
            "Saved batch: %s (size=%d)",
            batch_record["batch_sim_title"],
            len(batch_record["batch_sim_res_collection"])
        )
    except Exception as e:
        current_app.logger.exception("[simulate_batch_mp_end] ERROR: %s", e)

    status = "completed" if not errors else "completed_with_errors"
    return jsonify({
        "status": status,
        "total": len(combos),
        "ok": len(results),
        "errors": errors,
    })

@bp.route("/export_an_excel_file", methods=["POST"])
def export_as_excel():
    """POST JSON:
       { "batch_name": "<exact batch_sim_title>" }

       Saves an Excel file beside res_db.json and returns its filename and path.
    """
    print("Export Starting ")
    try:
        body = request.get_json(force=True, silent=False) or {}
        batch_name = body.get("batch_name")
        if not batch_name:
            return jsonify({"status": "error", "error": "missing batch_name"}), 400

        # Locate DB
        db_path = Path(current_app.instance_path) / "res_db.json"
        if not db_path.exists():
            return jsonify({"status": "error", "error": "db_non_exist"}), 200
        print("DB Located")
        try:
            data = json.loads(db_path.read_text(encoding="utf-8"))
        except Exception as e:
            current_app.logger.exception("Failed to read/parse res_db.json")
            return jsonify({"status": "error", "error": "db_read_failed", "detail": str(e)}), 500

        if not isinstance(data, list) or not data:
            return jsonify({"status": "error", "error": "db_empty"}), 200

        # Find the batch by title
        batch = next((b for b in data if b.get("batch_sim_title") == batch_name), None)
        if not batch:
            return jsonify({"status": "error", "error": "batch_not_found"}), 404

        # Build workbook
        wb = Workbook()
        # openpyxl creates a default sheet; we'll reuse it as Batch_Summary
        ws_summary = wb.active
        ws_summary.title = "Batch_Summary"
        print("Batch Summary Build Start ")
        # Batch summary (key/value style)
        row = 1
        _write_key_value_row(ws_summary, row, "batch_sim_title", batch.get("batch_sim_title", "")); row += 1
        _write_key_value_row(ws_summary, row, "batch_sim_time_stamp", batch.get("batch_sim_time_stamp", "")); row += 1
        _write_key_value_row(ws_summary, row, "vessel_name", batch.get("vessel_name", "")); row += 1
        _write_key_value_row(ws_summary, row, "batch_size", batch.get("batch_size", "")); row += 1
        _write_key_value_row(ws_summary, row, "actual_iterations", len(batch.get("batch_sim_res_collection", []) or [])); row += 1
        _autosize_columns(ws_summary)
        print("Batch Summary Build Ends")
        # Iterations_Summary (immediately after)
        ws_iter_summary = wb.create_sheet(title="Iterations_Summary", index=1)
        iterations = batch.get("batch_sim_res_collection", []) or []
        print("Iteration Summary Build Starts")
        # Create scalar rows
        scalar_rows = []
        for idx, it in enumerate(iterations):
            series, scalars = _split_iteration_fields(it)
            row_obj = {"iteration_idx": idx}
            row_obj.update(scalars)
            scalar_rows.append(row_obj)
        
        # Write Iterations_Summary as table
        if scalar_rows:
            headers = list(scalar_rows[0].keys())
            ws_iter_summary.append(headers)
            for r in scalar_rows:
                ws_iter_summary.append([r.get(h) for h in headers])
        else:
            ws_iter_summary.append(["note"])
            ws_iter_summary.append(["No iterations"])
        _autosize_columns(ws_iter_summary)
        print("Iteration Summary Build Ends")
        print("Iteration Sheets Build Starts")
        # Iteration sheets (one per iteration), start from index=2 to keep ordering
        used_names = set()
        for idx, it in enumerate(iterations):
            series, _scalars = _split_iteration_fields(it)

            # Base name: prefer sim_name; otherwise D/M/B counts
            base = idx
            if not base:
                d = it.get("die_count", it.get("Diesel_Engine_Count", "?"))
                m = it.get("meth_count", it.get("Meth_Count", "?"))
                b = it.get("bat_count", it.get("Battery_Count", "?"))
                base = "00"
                #base = f"D{d}_M{m}_B{b}"

            base = _sanitize_sheet_name(base)

            name = base
            n = 1
            while name in used_names or not name:
                suffix = f"_{n}"
                head = base[: MAX_SHEET_NAME_LEN - len(suffix)]
                name = _sanitize_sheet_name(head + suffix)
                n += 1
            used_names.add(name)

            ws = wb.create_sheet(title=name)
            rows = _series_to_rows(series)
            for r in rows:
                ws.append(r)
            _autosize_columns(ws)
        print("Iteration Sheets Build Ends")
        # Save beside the DB
        print("Saving Starts")
        safe_title = _sanitize_sheet_name(batch.get("batch_sim_title") or "batch")
        filename = f"{safe_title}.xlsx"
        out_path = db_path.parent / filename
        wb.save(out_path)
        print("Saving Done")
        # If you want to send the file directly, use:
        # return send_file(out_path, as_attachment=True, download_name=filename)

        return jsonify({
            "status": "ok",
            "excel_filename": filename,
            "excel_path": str(out_path)
        }), 200

    except Exception as e:
        current_app.logger.exception("export_as_excel failed")
        return jsonify({"status": "error", "error": str(e)}), 500
@bp.route("/return_all_batch_names", methods=["GET"])
def retrive_allbatch_names():
    # Link DataBase 
    db_path = Path(current_app.instance_path) / "res_db.json"
    db_non_exist_payload = "db_non_exist"

    # Check if database exist 
    if not db_path.exists():
        return jsonify({"result": db_non_exist_payload})

    # read into the database
    try:
        with open(db_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list) or len(data) == 0:
            return jsonify({"result": db_non_exist_payload})

        # Extract batch_sim_title from each batch entry
        batch_name_array = [batch.get("batch_sim_title", "unknown") for batch in data]
        print(batch_name_array)
        return jsonify({"result": batch_name_array})

    except Exception as e:
        current_app.logger.error(f"[return_all_batch_names] ERROR: {e}")
        return jsonify({"result": db_non_exist_payload})

@bp.route("/return_best_co2_cases", methods=["GET"])
def retrive_best_co2_cases():
    try:
        # Receive Batch Name from front end
        batch_name = request.args.get("batch_name", type=str)
        if not batch_name:
            current_app.logger.warning("[return_best_co2_cases] missing batch_name")
            return jsonify({"result": "missing_batch_name"}), 400

        # Attemps to search JSON database Path 
        db_path = Path(current_app.instance_path) / "res_db.json"
        if not db_path.exists():
            return jsonify({"result": "db_non_exist"}), 200

        # Attemp to load the database content
        try:
            data = json.loads(db_path.read_text(encoding="utf-8"))
        except Exception as e:
            current_app.logger.error("[return_best_co2_cases] JSON read error: %s", e)
            return jsonify({"result": "db_non_exist"}), 200

        # In case database is empty or the data is not a list 
        if not isinstance(data, list) or not data:
            return jsonify({"result": "db_non_exist"}), 200

        # Find the batch by exact title, target batch
        batch = next((b for b in data if b.get("batch_sim_title") == batch_name), None)
        if not batch:
            current_app.logger.info("[return_best_co2_cases] batch not found: %s", batch_name)
            return jsonify({"result": "batch_not_found"}), 200

        # Find  all iteration within the batch 
        iters = batch.get("batch_sim_res_collection", []) or []

        # Helper function to parse float 
        def to_float(x, default=0.0):
            try:
                if x is None:
                    return default
                return float(x)
            except Exception:
                return default

        # Helper to convert to int
        def to_int(x, default=0):
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return default

        # Filter: capable of meeting peak demand AND diesel-only (meth_count == 0)
        filtered = []
        
        for it in iters:
            max_pwr_potential_kw = to_float(it.get("max_pwr_potential"), 0.0)   # kW
            peak_power_demand_w  = to_float(it.get("peak_power_demand"), 0.0)   # W
            max_potential_w      = max_pwr_potential_kw * 1000.0                # temp conversion
            meth_count = to_int(it.get("meth_count", it.get("Meth_Count", 0)), 0)
            if max_potential_w >= peak_power_demand_w and meth_count == 0:
                filtered.append({
                    "iteration_id": to_int(it.get("iteration_id"), -1),
                    "sim_name": it.get("sim_name", ""),
                    "diesel_gen_name": it.get("diesel_gen_name", ""),
                    "diesel_count": to_int(it.get("diesel_count", it.get("die_count", 0)), 0),
                    "meth_gen_name": it.get("meth_gen_name", ""),
                    "meth_count": meth_count,
                    "bat_name": it.get("bat_name", ""),
                    "bat_count": to_int(it.get("bat_count", 0), 0),
                    "diesel_usage" : to_float(it.get("diesel_usage (Ton)"), 0.0),
                    "meth_usage"  : to_float(it.get("meth_usage (Ton)"), 0.0),
                    "cost" : to_float(it.get("captital_cost(£)"), 0.0),  
                    "CO2_emission": to_float(it.get("CO2_emission (Ton)"), float("inf")),
                    "penalty": to_float(it.get("penalty (EUR)"), 0.0),
                    "peak_power_demand": peak_power_demand_w,
                    "max_pwr_potential_w": max_potential_w,
                    "time" : it.get("time (h)", [1,2,3]),
                    "power_demand" : it.get("power_demand (KW)", [1,2,3]),
                    "gen_1_power" : it.get("gen_1_power (KW)",[1,2,3]),
                    "gen_2_power" : it.get("gen_2_power (KW)",[1,2,3]),
                    "gen_3_power" : it.get("gen_3_power (KW)",[1,2,3]),
                    "alt_gen_1_power" : it.get("alt_gen_1_power (KW)", [1,2,3]),
                    "alt_gen_2_power" : it.get("alt_gen_2_power (KW)", [1,2,3]),
                    "alt_gen_3_power" : it.get("alt_gen_3_power (KW)", [1,2,3]),
                    "bat_power" : it.get("battery_discharge (W)",[1,2,3]),
                    "node_tracker":it.get("node_tracker", [1,2,3])
                })

        # Sort ascending by CO2 and take up to 3
        filtered.sort(key=lambda x: x["CO2_emission"])
        top = filtered[:3]

        return jsonify({"result": top}), 200

    except Exception as e:
        current_app.logger.exception("[return_best_co2_cases] unexpected error: %s", e)
        # Ensure we *always* return something
        return jsonify({"result": "internal_error", "detail": str(e)}), 500

@bp.route("/return_best_diesel_consump_cases", methods=["GET"])
def retrive_best_diesel_consump_cases():
    try:
        # Retrive target batch name fomr frontend 
        print("received request")
        batch_name = request.args.get("batch_name", type = str)
        if not  batch_name: 
            current_app.logger.warning("[return best diesel case] missing batch_name")
            return jsonify({"result": "missing_batch_name"}), 400
        print("linking database")
        # Attemps to search JSON database path 
        db_path = Path(current_app.instance_path) / "res_db.json"
        if not db_path.exists():
            return jsonify({"result": "db_non_exist"}), 200

        # Attemps to load the database content
        try:
            print("loading database")
            data = json.loads(db_path.read_text(encoding="utf-8"))
        except Exception as e:
            print("[return best diesel case] JSON read error: %s", e)
            return jsonify({"result": "db_non_exist"}), 200
        
        # In case database is empty or the data is not a list 
        if not isinstance(data, list) or not data:
            print("database is empty")
            return jsonify({"result": "db_non_exist"}), 200

        # Find the batch by exact title, target batch
        print("finding batch title")
        batch = next((b for b in data if b.get("batch_sim_title") == batch_name), None)
        if not batch:
            print("[return best diesel case] batch not found: %s", batch_name)
            return jsonify({"result": "batch_not_found"}), 200
        
        # Find  all iteration within the batch 
        iters = batch.get("batch_sim_res_collection", []) or []

        # Helper function to parse float 
        def to_float(x, default=0.0):
            try:
                if x is None:
                    return default
                return float(x)
            except Exception:
                return default

        # Helper to convert to int
        def to_int(x, default=0):
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return default

        # Filter: capable of meeting peak demand AND diesel-only (meth_count == 0)
        filtered = []
        print("filtering  list")
        for it in iters:
            max_pwr_potential_kw = to_float(it.get("max_pwr_potential"), 0.0)   # kW
            peak_power_demand_w  = to_float(it.get("peak_power_demand"), 0.0)   # W
            max_potential_w      = max_pwr_potential_kw * 1000.0                # temp conversion

            meth_count = to_int(it.get("meth_count", it.get("Meth_Count", 0)), 0)
            if max_potential_w >= peak_power_demand_w and meth_count == 0:
                filtered.append({
                    "iteration_id": to_int(it.get("iteration_id"), -1),
                    "sim_name": it.get("sim_name", ""),
                    "diesel_gen_name": it.get("diesel_gen_name", ""),
                    "diesel_count": to_int(it.get("diesel_count", it.get("die_count", 0)), 0),
                    "meth_gen_name": it.get("meth_gen_name", ""),
                    "meth_count": meth_count,
                    "bat_name": it.get("bat_name", ""),
                    "bat_count": to_int(it.get("bat_count", 0), 0),
                    "diesel_usage" : to_float(it.get("diesel_usage (Ton)"), 0.0),
                    "meth_usage"  : to_float(it.get("meth_usage (Ton)"), 0.0),
                    "cost" : to_float(it.get("captital_cost(£)"), 0.0),  
                    "CO2_emission": to_float(it.get("CO2_emission (Ton)"), float("inf")),
                    "penalty": to_float(it.get("penalty (EUR)"), 0.0),
                    "peak_power_demand": peak_power_demand_w,
                    "max_pwr_potential_w": max_potential_w,
                    "time" : it.get("time (h)", [1,2,3]),
                    "power_demand" : it.get("power_demand (KW)", [1,2,3]),
                    "gen_1_power" : it.get("gen_1_power (KW)",[1,2,3]),
                    "gen_2_power" : it.get("gen_2_power (KW)",[1,2,3]),
                    "gen_3_power" : it.get("gen_3_power (KW)",[1,2,3]),
                    "alt_gen_1_power" : it.get("alt_gen_1_power (KW)", [1,2,3]),
                    "alt_gen_2_power" : it.get("alt_gen_2_power (KW)", [1,2,3]),
                    "alt_gen_3_power" : it.get("alt_gen_3_power (KW)", [1,2,3]),
                    "bat_power" : it.get("battery_discharge (W)",[1,2,3]),
                    "node_tracker":it.get("node_tracker", [1,2,3])
                })

        # Sort ascending by diesel usage and take up to 3
        filtered.sort(key=lambda x: x["diesel_usage"])
        top = filtered[:3]

        print("process end")
        return jsonify({"result": top}), 200

    except Exception as e:
        current_app.logger.exception("[return_best_co2_cases] unexpected error: %s", e)
        # Ensure we *always* return something
        return jsonify({"result": "internal_error", "detail": str(e)}), 500
    
@bp.route("/return_best_penalty_cases", methods =["GET"])
def retrive_best_penalty_cases():
    try:
        # Retrive target batch name fomr frontend 
        batch_name = request.args.get("batch_name", type = str)
        if not  batch_name: 
            current_app.logger.warning("[return_best_co2_cases] missing batch_name")
            return jsonify({"result": "missing_batch_name"}), 400

        # Attemps to search JSON database path 
        db_path = Path(current_app.instance_path) / "res_db.json"
        if not db_path.exists():
            return jsonify({"result": "db_non_exist"}), 200

        # Attemps to load the database content
        try:
            data = json.loads(db_path.read_text(encoding="utf-8"))
        except Exception as e:
            current_app.logger.error("[return_best_co2_cases] JSON read error: %s", e)
            return jsonify({"result": "db_non_exist"}), 200
        
        # In case database is empty or the data is not a list 
        if not isinstance(data, list) or not data:
            return jsonify({"result": "db_non_exist"}), 200

        # Find the batch by exact title, target batch
        batch = next((b for b in data if b.get("batch_sim_title") == batch_name), None)
        if not batch:
            current_app.logger.info("[return_best_co2_cases] batch not found: %s", batch_name)
            return jsonify({"result": "batch_not_found"}), 200
        
        # Find  all iteration within the batch 
        iters = batch.get("batch_sim_res_collection", []) or []

        # Helper function to parse float 
        def to_float(x, default=0.0):
            try:
                if x is None:
                    return default
                return float(x)
            except Exception:
                return default

        # Helper to convert to int
        def to_int(x, default=0):
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return default

        # Filter: capable of meeting peak demand AND diesel-only (meth_count == 0)
        filtered = []

        # Filter Process Begine
        for it in iters:
            # KW to W conversion for consistancy 
            max_pwr_potential_kw = to_float(it.get("max_pwr_potential"), 0.0)   
            peak_power_demand_w  = to_float(it.get("peak_power_demand"), 0.0)   
            max_potential_w      = max_pwr_potential_kw * 1000.0               

            meth_count = to_int(it.get("meth_count", it.get("Meth_Count", 0)), 0)
            
            # Filter thouse who can not meet the demand and use a methanol gen
            if max_potential_w >= peak_power_demand_w and meth_count == 0:
                filtered.append({
                    "iteration_id": to_int(it.get("iteration_id"), -1),
                    "sim_name": it.get("sim_name", ""),
                    "diesel_gen_name": it.get("diesel_gen_name", ""),
                    "diesel_count": to_int(it.get("diesel_count", it.get("die_count", 0)), 0),
                    "meth_gen_name": it.get("meth_gen_name", ""),
                    "meth_count": meth_count,
                    "bat_name": it.get("bat_name", ""),
                    "bat_count": to_int(it.get("bat_count", 0), 0),
                    "diesel_usage" : to_float(it.get("diesel_usage (Ton)"), 0.0),
                    "meth_usage"  : to_float(it.get("meth_usage (Ton)"), 0.0),
                    "cost" : to_float(it.get("captital_cost(£)"), 0.0),  
                    "CO2_emission": to_float(it.get("CO2_emission (Ton)"), float("inf")),
                    "penalty": to_float(it.get("penalty (EUR)"), 0.0),
                    "peak_power_demand": peak_power_demand_w,
                    "max_pwr_potential_w": max_potential_w,
                    "time" : it.get("time (h)", [1,2,3]),
                    "power_demand" : it.get("power_demand (KW)", [1,2,3]),
                    "gen_1_power" : it.get("gen_1_power (KW)",[1,2,3]),
                    "gen_2_power" : it.get("gen_2_power (KW)",[1,2,3]),
                    "gen_3_power" : it.get("gen_3_power (KW)",[1,2,3]),
                    "alt_gen_1_power" : it.get("alt_gen_1_power (KW)", [1,2,3]),
                    "alt_gen_2_power" : it.get("alt_gen_2_power (KW)", [1,2,3]),
                    "alt_gen_3_power" : it.get("alt_gen_3_power (KW)", [1,2,3]),
                    "bat_power" : it.get("battery_discharge (W)",[1,2,3]),
                    "node_tracker":it.get("node_tracker", [1,2,3])
                })

        # Sort ascending by penalty and take up to 3
        filtered.sort(key=lambda x: x["penalty"])
        top = filtered[:3]

        return jsonify({"result": top}), 200

    except Exception as e:
        current_app.logger.exception("[return_best_co2_cases] unexpected error: %s", e)
        # Ensure we *always* return something
        return jsonify({"result": "internal_error", "detail": str(e)}), 500
    
@bp.route("/simulate_batch/status", methods=["GET"])
def simulate_batch_status():
    # returns current in-memory status so frontend can poll
    return jsonify(PROGRESS)

@bp.route("/simulate_batch/cancel", methods=["POST"])
def simulate_batch_cancel():
    # flag the current batch to stop after current simulation
    PROGRESS["cancelled"] = True
    return jsonify({"status": "ok", "message": "cancellation requested"})

@bp.route("/get_class_names", methods=["POST"])
def get_class_names():
    omc = OMCConnection()
    omc.request(
        f"loadFile(\"{current_app.instance_path}/{request.json['model_name']}.mo\")"
    )
    names = omc.request("getClassNames()")
    return json.dumps([n for n in names[1:-2].split(",")])

@bp.route("/get_parameter_names", methods=["POST"])
def get_parameter_names():
    omc = OMCConnection()
    omc.request(
        f"loadFile(\"{current_app.instance_path}/{request.json['model_name']}.mo\")"
    )
    params = omc.request(f"getParameterNames({request.json['class']})")

    return json.dumps([n for n in params[1:-2].split(",")])

@bp.route("/set_duty_cycle", methods=["POST"])
def set_duty_cycle():
    # decode the duty cycle from base64 and save it to a file
    try:
        with current_app.open_instance_resource(
                f"{request.json['model_name']}_duty_cycle.txt", "w") as duty_cycle_file:
            duty_cycle_file.write(
                base64.b64decode(request.json['duty_cycle_data']).decode("utf-8"))
            
            duty_cycle_path = f"{current_app.instance_path}/{request.json['model_name']}_duty_cycle.txt"
            print(f"Decoded duty cycle to {duty_cycle_path}")
            
            model_path = f"{current_app.instance_path}/{request.json['model_name']}.mo"
            # Check if the model path exists
            if os.path.exists(model_path):
                print(f"Model file found: {model_path}")
                with open(model_path, "r") as f:
                    lines = f.readlines()

                new_lines = []
                replaced = False
                for line in lines:
                    if not replaced and "Modelica.Blocks.Tables.CombiTable1Ds" in line and "fileName" in line:
                        # Update the duty cycle path
                        new_line = line.replace(
                            re.search(r'fileName\s*=\s*".*?"', line).group(0),
                            f'fileName = "{duty_cycle_path}"'
                        )
                        new_lines.append(new_line)
                        replaced = True
                    else:
                        new_lines.append(line)

                with open(model_path, "w") as f:
                    f.writelines(new_lines)
                    print(f"Replaced duty cycle path in {model_path} with {duty_cycle_path}")
            else:
                print(f"Model path does not exist: {model_path}")
    except OSError:
        return json.dumps({
            "name": request.json['model_name'],
            "status": "Error opening writable instance resource"
        })

    return json.dumps({
        "name": request.json['model_name'],
        "status": "Duty cycle written"
    })


### Helper Functions ### 
def get_db_path() -> Path:
    # instance_path is the right place to write app data
    p = Path(current_app.instance_path) / "res_db.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p
def read_db(db_path: Path):
    try:
        text = db_path.read_text(encoding="utf-8")
        data = json.loads(text)
        if not isinstance(data, list):
            current_app.logger.warning("res_db.json not a list; reinitializing")
            return []
        return data
    except FileNotFoundError:
        return []
    except Exception as e:
        current_app.logger.exception("Failed to read res_db.json; reinitializing")
        return []
def write_db_atomic(db_path: Path, data):
    tmp = db_path.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(db_path)  
def append_batch_result(batch_record: dict):
    db_path = get_db_path()
    data = read_db(db_path)
    data.append(batch_record)
    write_db_atomic(db_path, data)

def get_available_batches():
    db_path = Path(current_app.instance_path) / "res_db.json"
    if not db_path.exists():
        return []

    try:
        with open(db_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract batch names if present
        batch_names = [batch.get("batch_sim_title", "Unnamed Batch") for batch in data]
        return batch_names
    except Exception as e:
        print(f"[get_available_batches] ERROR: {e}")
        return []
def single_simulation(model_name: str,
                      start_time: float,
                      stop_time: float,
                      index: int,
                      override_configuration: dict,
                      instance_path: str) -> dict:
    """
    Runs one simulation in its own process, writes result under a unique folder,
    reads CSV, processes, and RETURNS a JSON-serializable dict.
    """
    # Unique output dir per job (prevents accessing the same file)
    job_dir = Path(instance_path) / "runs" / f"{model_name}_job_{index}"
    job_dir.mkdir(parents=True, exist_ok=True)

    # Connect OMC *inside* the process
    omc = OMCConnection()
    # Use the instance_path passed in
    omc.request(f'loadFile("{instance_path}/{model_name}.mo")')

    overrides_list = override_configuration.get("changed_parameters", [])
    ov = " -override " + ",".join(f"{o['param']}={o['value']}" for o in overrides_list) if overrides_list else ""
    simflags = f'-outputPath {job_dir}{ov}'

    # Run simulate
    omc.request(
        f'simulate({model_name}, '
        f'outputFormat="csv", '
        f'startTime={start_time}, '
        f'stopTime={stop_time}, '
        f'simflags="{simflags}")'
    )

    # Read the CSV from *this* job’s folder
    csv_path = job_dir / f"{model_name}_res.csv"
    # process CSV → result dict (must be JSON-serializable)
    inst = override_configuration.get("instance", {})
    d_name = inst.get("Diesel_Engine", {}).get("engine_name", "")
    m_name = inst.get("Meth_Engine", {}).get("engine_name", "")
    b_name = inst.get("Battery", {}).get("battery_name", "")
    d_cnt  = int(inst.get("Diesel_Engine_Count", 0) or 0)
    m_cnt  = int(inst.get("Meth_Count", 0) or 0)
    b_cnt  = int(inst.get("Battery_Count", 0) or 0)

    sim_name = f"{d_cnt}*{d_name} + {m_cnt}*{m_name} + {b_cnt}*{b_name}"

    # Use robust processing Chat GPT said this will handle parsing the csv data? 
    res = process_simulation_result_safe(
        index=index,
        simName=sim_name,
        dieselGenName=d_name,
        dieselGenCount=d_cnt,
        methGenName=m_name,
        methGenCount=m_cnt,
        batName=b_name,
        batCount=b_cnt,
        csv_path=str(csv_path),
    )
    # Remove the modelica Simulation output file after the content of which is read 
    try:
        shutil.rmtree(job_dir)
        print(f"[simulate_batch] cleaned up {job_dir}")
    except Exception as e:
        print(f"[simulate_batch] cleanup failed for {job_dir}: {e}")
    return res
def process_simulation_result_safe(index, simName, dieselGenName, dieselGenCount,
                                   methGenName, methGenCount, batName, batCount,
                                   csv_path):
  
    def last_non_null(df, col, default=0):
        if col not in df.columns: return default
        s = df[col].dropna()
        if s.empty: return default
        v = s.iloc[-1]
        if isinstance(v, np.generic): return v.item()
        return float(v) if isinstance(v, (int, float, np.number)) else v

    try:
        df = pd.read_csv(csv_path)
    except Exception:
        df = pd.DataFrame()

    def col_or_empty(name):
        return df[name].tolist() if name in df.columns else []

    # Calculate the EU certified Emission before returning the result 
    total_emission = 0 
    penalty  = 0
    try :
        temp_emission_res = calculate_eu_fuel_compliance(last_non_null(df, "realValueFuelMass.showNumber", 0),
                                                         last_non_null(df, "realValueAltFuelMass.showNumber", 0))
        total_emission = temp_emission_res.get("emission").get("total_co2")
        penalty = temp_emission_res.get("emission").get("penalty")
    except Exception as e :
        print(f"Total Emission: {total_emission} + Total Penalty :{penalty}")

    # Form the result packcage
    out = {
        "sim_name": simName,
        "diesel_gen_name": dieselGenName,
        "meth_gen_name": methGenName,
        "bat_name": batName,
        "diesel_count": int(dieselGenCount),
        "meth_count": int(methGenCount),
        "bat_count": int(batCount),
        "iteration_id": int(index),
        "time (s)": col_or_empty("time"),
        "power_demand (W)": col_or_empty("gain1.y"),
        "gen_1_power (W)": col_or_empty("generator1.P_out"),
        "gen_2_power (W)" : col_or_empty("generator2.P_out"),
        "gen_3_power (W)" : col_or_empty("generator3.P_out"),
        "gen_1_fuel_mass_flow (g/s)" : col_or_empty("generator1.m_flow_fuel"),
        "gen_2_fuel_mass_flow (g/s)" : col_or_empty("generator2.m_flow_fuel"),
        "gen_3_fuel_mass_flow (g/s)" : col_or_empty("generator3.m_flow_fuel"),
        "gen_1_fuel_volume_flow (m^3/s)" : col_or_empty("generator1.V_flow_fuel"),
        "gen_2_fuel_volume_flow (m^3/s)" : col_or_empty("generator2.V_flow_fuel"),
        "gen_3_fuel_volume_flow (m^3/s)" : col_or_empty("generator3.V_flow_fuel"),
        "alt_gen_1_power (W)":col_or_empty("generator_af_1.P_out"),
        "alt_gen_2_power (W)":col_or_empty("generator_af_2.P_out"),
        "alt_gen_3_power (W)":col_or_empty("generator_af_3.P_out"),
        "alt_gen_1_fuel_mass_flow (g/s)" : col_or_empty("generator_af_1.m_flow_fuel"),
        "alt_gen_2_fuel_mass_flow (g/s)" : col_or_empty("generator_af_2.m_flow_fuel"),
        "alt_gen_3_fuel_mass_flow (g/s)" : col_or_empty("generator_af_3.m_flow_fuel"),
        "alt_gen_1_fuel_volume_flow (m^3/s)" :col_or_empty("generator_af_1.V_flow_fuel"),
        "alt_gen_2_fuel_volume_flow (m^3/s)" :col_or_empty("generator_af_2.V_flow_fuel"),
        "alt_gen_3_fuel_volume_flow (m^3/s)" :col_or_empty("generator_af_3.V_flow_fuel"),
        "battery_soc (%)": col_or_empty("battery1.SOC"),
        "battery_discharge (W)": col_or_empty("battery1.P_discharge"),
        "diesel_usage (KG)": last_non_null(df, "realValueFuelMass.showNumber", 0),
        "methanol_usage (KG)": last_non_null(df, "realValueAltFuelMass.showNumber", 0),
        "CO2_emission": float(total_emission), 
        "penalty": float(penalty),       
    }
    return out
def calculate_eu_fuel_compliance(dieselConsumption, methanolConsumption):
    
    fuel_type_mapping  = {
        'Diesel' : 'MDO MGO (Grades DMX to DMB)',
        'Methanol' : 'e-methanol E10'
    }
    fuel_results = {}
    if  dieselConsumption  > 0 :
        fuel_results['Diesel'] = dieselConsumption
    if  methanolConsumption > 0 :
        fuel_results['Methanol'] = methanolConsumption
    emission_calculator = EUMarineEmissionCalculator()    
    
    emission_fuel_data = [] # Contains  fuel invoved in CO2 PRoduction On the vessel  
    total_emission_tco2eq = 0
    emission_breakdown = {}

    for fuel_type, consumption_kg in fuel_results.items():
        if consumption_kg > 0:
            # Convert kg to tonnes for emission calculation
            consumption_tonnes = consumption_kg / 1000
            
            emission_fuel_type = fuel_type_mapping.get(fuel_type)
            if not emission_fuel_type:
                print(f"No emission fuel mapping found for {fuel_type}")
                continue
            
            # All fuel consumption is within EU/EEA
            fuel_data_entry = {
                'fuel_type': emission_fuel_type,
                'fuel_consumption_within_eu': consumption_tonnes,
                'fuel_consumption_in_out_eu': 0.0,  # All within EU
                'is_biofuel': False,
                'biofuel_option': 3,
                'biofuel_percentage': 0,
                'biofuel_pathway': ''
            }
            
            emission_fuel_data.append(fuel_data_entry)
            print(f"{fuel_type} ({emission_fuel_type}): {consumption_tonnes:.3f} tonnes within EU")
    
    # Calculate emissions if all data is presented
    emission_results = None
    compliance_results = None
    if emission_fuel_data:
        try:
            # Validate fuel data
            validation = emission_calculator.validate_fuel_data(emission_fuel_data)
            if not validation['valid']:
                print(f"Invalid fuel data for emission calculation: {validation['error']}")
                return 
            
            # Run Phase 2 emission calculation (includes both phase 1 and 2)
            emission_results = emission_calculator.calculate_emissions_phase2(emission_fuel_data, target_year=2025)
            
            # Extract emission breakdown per fuel
            fuel_breakdown = emission_results.get('fuel_breakdown', [])
            for fuel_calc in fuel_breakdown:
                original_fuel_type = None
                emission_fuel_type = fuel_calc.get('fuel_type', '')
                
                # Reverse map to original fuel type
                for orig_type, mapped_type in fuel_type_mapping.items():
                    if mapped_type == emission_fuel_type:
                        original_fuel_type = orig_type
                        break
                
                if original_fuel_type:
                    emission_breakdown[original_fuel_type] = {
                        'consumption_tonnes': fuel_calc.get('fuel_mass_total', 0),
                        'ghg_emission_tonnes': fuel_calc.get('ghg_emission', 0)
                    }
            
            # Extract compliance results
            final_results = emission_results.get('final_results', {})
            compliance_results = {
                'ghg_intensity_actual': final_results.get('ghg_intensity_actual', 0),
                'ghg_intensity_target': final_results.get('ghg_intensity_target_2025_2029', 0),
                'compliance_balance_tco2eq': final_results.get('compliance_balance_tco2eq', 0),
                'penalty_eur': final_results.get('penalty', 0),
                'total_ghg_emission_tonnes': final_results.get('ghg_emission', 0)
            }
            
            total_emission_tco2eq = final_results.get('ghg_emission', 0)
            
            print(f"Emission calculation completed:")
            print(f"---Total GHG Emission: {total_emission_tco2eq:.3f} tCO2eq")
            print(f"---GHG Intensity Actual: {compliance_results['ghg_intensity_actual']:.6f} gCO2eq/MJ")
            print(f"---GHG Intensity Target: {compliance_results['ghg_intensity_target']:.6f} gCO2eq/MJ")
            print(f"---Compliance Balance: {compliance_results['compliance_balance_tco2eq']:.6f} tCO2eq")
            print(f"---Penalty: EUR {compliance_results['penalty_eur']:.2f}")
            
        except Exception as e:
            print(f"Emission calculation failed: {str(e)}")
            # Continue with fuel consumption results but no emissions
            print("Continuing with fuel consumption results only")
    return  {
        "emission": {
            "total_co2": compliance_results.get('total_ghg_emission_tonnes', 555),
            "penalty" : compliance_results.get('penalty_eur', 555)
        }
    }
def process_simmultion_result (index, simName, dieselGenName, dieselGenCount, methGenName, methGenCount, batName, batCount, total_cost, max_powertrain_gen):
    processed_simulation_result = {}
    try:
        df = pd.read_csv('./instance/AZEAT_FortunaCrane_Valid_res.csv')
        processed_simulation_result ['sim_name'] = simName
        processed_simulation_result ['diesel_gen_name'] = dieselGenName
        processed_simulation_result ['meth_gen_name'] = methGenName
        processed_simulation_result ['bat_name'] = batName
        processed_simulation_result ['die_count'] = dieselGenCount
        processed_simulation_result ['meth_count'] = methGenCount
        processed_simulation_result ['bat_count'] = batCount
        processed_simulation_result ['iteration_id'] = index
        processed_simulation_result ['time (h)'] = [t / 3600 for t in df['time'].tolist()]
        processed_simulation_result ['power_demand (KW)'] = [p / 1000 for p in df['gain1.y'].tolist()]
        processed_simulation_result ['gen_1_power (KW)'] = [p / 1000 for p in df['generator1.P_out'].tolist()]
        processed_simulation_result ['gen_2_power (KW)'] = [p / 1000 for p in df['generator2.P_out'].tolist()]
        processed_simulation_result ['gen_3_power (KW)'] = [p / 1000 for p in df['generator3.P_out'].tolist()]
        processed_simulation_result ['gen_1_fuel_mass_flow (g/s)'] = df['generator1.m_flow_fuel'].tolist()
        processed_simulation_result ['gen_2_fuel_mass_flow (g/s)'] = df['generator2.m_flow_fuel'].tolist()
        processed_simulation_result ['gen_3_fuel_mass_flow (g/s)'] = df['generator3.m_flow_fuel'].tolist()
        processed_simulation_result ['gen_1_fuel_volume_flow (m^3/s)'] = df['generator1.V_flow_fuel'].tolist()
        processed_simulation_result ['gen_2_fuel_volume_flow (m^3/s)'] = df['generator2.V_flow_fuel'].tolist()
        processed_simulation_result ['gen_3_fuel_volume_flow (m^3/s)'] = df['generator3.V_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_1_power (KW)'] = [p / 1000 for p in df['generator_af_1.P_out'].tolist()]
        processed_simulation_result ['alt_gen_2_power (KW)'] = [p / 1000 for p in df['generator_af_2.P_out'].tolist()]
        processed_simulation_result ['alt_gen_3_power (KW)'] = [p / 1000 for p in df['generator_af_3.P_out'].tolist()]
        processed_simulation_result ['alt_gen_1_fuel_mass_flow (g/s)'] = df['generator_af_1.m_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_2_fuel_mass_flow (g/s)'] = df['generator_af_2.m_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_3_fuel_mass_flow (g/s)'] = df['generator_af_3.m_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_1_fuel_volume_flow (m^3/s)'] = df['generator_af_1.V_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_2_fuel_volume_flow (m^3/s)'] = df['generator_af_2.V_flow_fuel'].tolist()
        processed_simulation_result ['alt_gen_3_fuel_volume_flow (m^3/s)'] = df['generator_af_3.V_flow_fuel'].tolist()
        processed_simulation_result ['battery_soc (%)'] = df['battery1.SOC'].tolist()
        processed_simulation_result ['battery_discharge (W)'] = (df['battery1.P_discharge_abs']).tolist()
        processed_simulation_result ['battery_charge (W)'] = (df['battery1.P_charge_abs']).tolist()
        processed_simulation_result ['diesel_usage (Ton)'] = float(df['realValueFuelMass.showNumber'].iloc[-1])/1000
        print(f"Diesel Comsumption: {df['realValueFuelMass.showNumber'].iloc[-1]} KG")
        processed_simulation_result ['meth_usage (Ton)'] = float(df['realValueAltFuelMass.showNumber'].iloc[-1])/1000
        print(f"Methanol Comsumption: {df['realValueAltFuelMass.showNumber'].iloc[-1]} KG" )
        print(f"Cost : {total_cost} £" )
        total_emisstion = 444
        penalty = 444
        try: 
            temp_emission_res = calculate_eu_fuel_compliance(df['realValueFuelMass.showNumber'].iloc[-1],df['realValueAltFuelMass.showNumber'].iloc[-1])
            total_emisstion = temp_emission_res.get("emission").get("total_co2")
            penalty = temp_emission_res.get("emission").get("penalty")
            print(f"Total Emission: {total_emisstion} + Total Penalty :{penalty}")
        except Exception as e:
            print(f"[EU Emission Calculator Does not Work] ERROR: {e}")
        processed_simulation_result ['CO2_emission (Ton)'] = total_emisstion
        processed_simulation_result ['penalty (EUR)'] = penalty
        processed_simulation_result ['captital_cost(£)'] = total_cost
        processed_simulation_result ['node_tracker'] = df['masterControllerSingleBattery.methanolBatteryDieselDebug.nodeTracker'].tolist()
        processed_simulation_result ['max_pwr_potential (KW)'] = max_powertrain_gen
        processed_simulation_result ['peak_power_demand (KW)'] = (max(df['gain1.y'].tolist())) * (1/1000)
    except Exception as e:
        print(f"[simulate_batch_result_processing] ERROR: {e}")
        return jsonify({"status": "error", "error": str(e)}), 500
    return processed_simulation_result 
 
def _sanitize_sheet_name(name: str) -> str:
    if not name:
        name = "Sheet"
    name = ILLEGAL_SHEET_CHARS.sub("", str(name)).strip()
    if len(name) > MAX_SHEET_NAME_LEN:
        name = name[:MAX_SHEET_NAME_LEN]
    return name or "Sheet"
def _is_listlike(v):
    return isinstance(v, list)
def _split_iteration_fields(iter_obj: dict):
    """Split a single iteration object into
       - series: dict of { key: list }
       - scalars: dict of non-list values
    """
    series = {}
    scalars = {}
    for k, v in (iter_obj or {}).items():
        if _is_listlike(v):
            series[k] = v
        else:
            scalars[k] = v
    return series, scalars
def _write_key_value_row(ws, row, key, value):
    ws.cell(row=row, column=1, value=key)
    ws.cell(row=row, column=2, value=value)
def _autosize_columns(ws):
    # Simple autosize by header and first 50 rows to avoid huge loops on large sheets
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        length = 0
        col_letter = get_column_letter(col_idx)
        # header + first ~50 rows
        for r in range(1, min(ws.max_row, 50) + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, length + 2), 50)
def _series_to_rows(series: dict):
    """Turn { col1: [...], col2: [...] } into [ [header...], [row1...], ... ]"""
    keys = list(series.keys())
    if not keys:
        return [["no_series_data"]]
    max_len = max(len(series.get(k, [])) for k in keys)
    rows = [keys]
    for i in range(max_len):
        rows.append([series.get(k, [None] * (i + 1))[i] if i < len(series.get(k, [])) else None for k in keys])
    return rows